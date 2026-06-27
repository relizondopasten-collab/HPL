"""Generación de reportes PDF / Excel para evaluaciones agrícolas."""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services import db
from app.services.anova import rcbd_anova_tukey
from app.services.cld import compact_letter_display
from app.services.efficacy import abbott, henderson_tilton

TOTAL_VARIABLE = "__total__"

TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"
_env = Environment(
    loader=FileSystemLoader(str(TEMPLATES_DIR)),
    autoescape=select_autoescape(["html", "xml"]),
)

STAGE_LABELS = {
    "egg": "huevo",
    "larva_n1": "L1",
    "larva_n2": "L2",
    "larva_n3": "L3",
    "larva_n4": "L4",
    "nymph": "ninfa",
    "pupa": "pupa",
    "adult": "adulto",
    "mobile_form": "móvil",
}


# =================== Carga + análisis ===================

def _build_dataset(
    plots: list[dict[str, Any]],
    counts: list[dict[str, Any]],
    variable: str,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Devuelve (observaciones, treatment_label_by_key).

    Cada parcela aporta UNA observación: el promedio de sus muestras para la
    variable elegida (estadio específico o suma total).
    """
    plot_meta: dict[str, dict[str, Any]] = {p["id"]: p for p in plots}

    samples_by_plot: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    has_sample: dict[str, set[int]] = defaultdict(set)

    for c in counts:
        plot_id = c["plot_id"]
        sample = c["sample_index"]
        stage = c.get("life_stage")
        alive = c.get("alive") or 0
        has_sample[plot_id].add(sample)
        if variable == TOTAL_VARIABLE:
            samples_by_plot[plot_id][sample] += alive
        elif stage == variable:
            samples_by_plot[plot_id][sample] = alive

    treatments = {p["treatment_id"] for p in plots}
    treatment_labels = _treatment_labels_by_key(plot_meta)

    observations: list[dict[str, Any]] = []
    for plot_id, samples in samples_by_plot.items():
        if not samples:
            continue
        plot = plot_meta.get(plot_id)
        if not plot:
            continue
        values = list(samples.values())
        observations.append(
            {
                "treatment": _treatment_key_for_plot(plot, plot_meta),
                "block": f"B{plot['block']}",
                "value": sum(values) / len(values),
            }
        )

    return observations, treatment_labels


def _treatment_key_for_plot(plot: dict[str, Any], _plot_meta: dict[str, Any]) -> str:
    return f"T{plot.get('_treatment_number')}"


def _treatment_labels_by_key(plot_meta: dict[str, dict[str, Any]]) -> dict[str, str]:
    out: dict[str, str] = {}
    for p in plot_meta.values():
        key = f"T{p.get('_treatment_number')}"
        out[key] = p.get("_treatment_label", "")
    return out


def _hydrate_plots(plots: list[dict[str, Any]], treatments: list[dict[str, Any]]) -> None:
    by_id = {t["id"]: t for t in treatments}
    for p in plots:
        t = by_id.get(p.get("treatment_id"))
        p["_treatment_number"] = t["number"] if t else None
        p["_treatment_label"] = t["label"] if t else ""
        p["_treatment_is_control"] = t["is_control"] if t else False


def _variable_label(variable: str) -> str:
    if variable == TOTAL_VARIABLE:
        return "Total de vivos (suma de estadios)"
    return STAGE_LABELS.get(variable, variable)


def analyze_evaluation(
    evaluation_id: str,
    variable: str = TOTAL_VARIABLE,
) -> dict[str, Any]:
    """Carga datos de Supabase, arma dataset y corre ANOVA + CLD + Abbott."""
    evaluation = db.get_evaluation(evaluation_id)
    if not evaluation:
        raise LookupError(f"Evaluación {evaluation_id} no encontrada")
    trial = db.get_trial(evaluation["trial_id"])
    if not trial:
        raise LookupError(f"Ensayo {evaluation['trial_id']} no encontrado")
    treatments = db.get_treatments(trial["id"])
    plots = db.get_plots(trial["id"])
    counts = db.get_pest_counts(evaluation_id)

    _hydrate_plots(plots, treatments)

    observations, treatment_labels = _build_dataset(plots, counts, variable)

    if len(observations) < 4:
        raise ValueError(
            "Se necesitan al menos 4 observaciones para análisis (≥2 tratamientos × ≥2 bloques)."
        )

    df = pd.DataFrame(observations)
    use_block = df["block"].notna().all() and df["block"].nunique() >= 2
    result = rcbd_anova_tukey(
        df,
        value_col="value",
        treatment_col="treatment",
        block_col="block" if use_block else None,
    )

    means_by_trt = {row["treatment"]: row["mean"] for row in result["means"]}
    differing = [(t["group1"], t["group2"]) for t in result["tukey"] if t["reject"]]
    letters = compact_letter_display(means_by_trt, differing)

    control_key = None
    for t in treatments:
        if t["is_control"]:
            control_key = f"T{t['number']}"
            break
    control_mean = means_by_trt.get(control_key) if control_key else None

    enriched = []
    for row in result["means"]:
        trt = row["treatment"]
        eff = (
            abbott(control_mean, row["mean"])
            if control_mean is not None and trt != control_key
            else (0.0 if trt == control_key else None)
        )
        enriched.append({**row, "letter": letters.get(trt, ""), "efficacy_pct": eff})
    enriched.sort(key=lambda r: -r["mean"])
    result["means"] = enriched
    result["control_treatment"] = control_key
    result["treatment_labels"] = treatment_labels
    result["observations"] = observations
    result["counts"] = counts
    result["plots"] = plots
    result["treatments"] = treatments
    result["evaluation"] = evaluation
    result["trial"] = trial
    result["variable"] = variable
    result["variable_label"] = _variable_label(variable)
    return result


# =================== PDF ===================

def render_evaluation_pdf(evaluation_id: str, variable: str = TOTAL_VARIABLE) -> bytes:
    """Renderiza el reporte PDF de una evaluación."""
    # Import diferido para que el módulo se pueda importar sin las libs nativas
    # de WeasyPrint instaladas (útil para tests del resto del backend).
    from weasyprint import HTML  # type: ignore

    data = analyze_evaluation(evaluation_id, variable)
    trial = data["trial"]
    evaluation = data["evaluation"]

    crop = db.get_crop(trial.get("crop_id"))
    client = db.get_client(trial.get("client_id"))
    pest = db.get_pest(evaluation.get("pest_id") or trial.get("pest_id"))

    max_mean = max((m["mean"] for m in data["means"]), default=0.0)

    html = _env.get_template("evaluation_report.html").render(
        trial=trial,
        evaluation=evaluation,
        treatments=data["treatments"],
        means=data["means"],
        anova=data["anova"],
        tukey=data["tukey"],
        cv_pct=data["cv_pct"],
        mse=data["mse"],
        n_obs=data["n_obs"],
        control=data["control_treatment"],
        treatment_labels=data["treatment_labels"],
        variable_label=data["variable_label"],
        max_mean=max_mean,
        crop=crop,
        client=client,
        pest=pest,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )

    return HTML(string=html).write_pdf()


# =================== Excel ===================

def render_evaluation_xlsx(evaluation_id: str, variable: str = TOTAL_VARIABLE) -> bytes:
    """Excel con resumen, dataset crudo y resultados."""
    data = analyze_evaluation(evaluation_id, variable)
    trial = data["trial"]
    evaluation = data["evaluation"]
    treatments = data["treatments"]
    plots = data["plots"]
    counts = data["counts"]

    wb = Workbook()

    # ---------- Hoja 1: Resumen ----------
    ws = wb.active
    ws.title = "Resumen"
    _heading(ws, "A1", f"{trial['code']} — {trial['name']}")
    rows = [
        ("Código", trial["code"]),
        ("Nombre", trial["name"]),
        ("Tipo", trial.get("trial_type")),
        ("Diseño", trial.get("design")),
        ("# Tratamientos", trial.get("n_treatments")),
        ("# Repeticiones", trial.get("n_replicates")),
        ("Ubicación", trial.get("location") or ""),
        ("Fecha evaluación", evaluation.get("evaluated_at")),
        ("DDA", evaluation.get("days_after_application") if evaluation.get("days_after_application") is not None else ""),
        ("Variable", data["variable_label"]),
        ("CV %", round(data["cv_pct"], 2)),
        ("MSE", round(data["mse"], 4)),
        ("n observaciones", data["n_obs"]),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = v
    _autosize(ws)

    # ---------- Hoja 2: Tratamientos ----------
    ws_t = wb.create_sheet("Tratamientos")
    headers = ["#", "Etiqueta", "Producto", "Dosis", "Unidad", "Testigo"]
    _header_row(ws_t, 1, headers)
    for i, t in enumerate(treatments, start=2):
        ws_t.cell(i, 1, f"T{t['number']}")
        ws_t.cell(i, 2, t.get("label"))
        ws_t.cell(i, 3, t.get("product_id"))
        ws_t.cell(i, 4, t.get("dose"))
        ws_t.cell(i, 5, t.get("dose_unit"))
        ws_t.cell(i, 6, "sí" if t.get("is_control") else "")
    _autosize(ws_t)

    # ---------- Hoja 3: Mapa de parcelas ----------
    ws_p = wb.create_sheet("Parcelas")
    _header_row(ws_p, 1, ["Parcela ID", "Bloque", "Col", "Tratamiento", "Etiqueta", "Testigo"])
    for i, p in enumerate(plots, start=2):
        ws_p.cell(i, 1, p["id"])
        ws_p.cell(i, 2, p["block"])
        ws_p.cell(i, 3, p.get("position_col"))
        ws_p.cell(i, 4, f"T{p['_treatment_number']}")
        ws_p.cell(i, 5, p.get("_treatment_label"))
        ws_p.cell(i, 6, "sí" if p.get("_treatment_is_control") else "")
    _autosize(ws_p)

    # ---------- Hoja 4: Datos crudos ----------
    ws_d = wb.create_sheet("Conteos")
    _header_row(
        ws_d,
        1,
        ["Bloque", "Col", "Tratamiento", "Muestra", "Estadio", "Vivos", "Muertos"],
    )
    plot_by_id = {p["id"]: p for p in plots}
    row = 2
    for c in counts:
        plot = plot_by_id.get(c["plot_id"])
        if not plot:
            continue
        ws_d.cell(row, 1, plot["block"])
        ws_d.cell(row, 2, plot.get("position_col"))
        ws_d.cell(row, 3, f"T{plot['_treatment_number']}")
        ws_d.cell(row, 4, c.get("sample_index"))
        ws_d.cell(row, 5, STAGE_LABELS.get(c.get("life_stage") or "", c.get("life_stage")))
        ws_d.cell(row, 6, c.get("alive") or 0)
        ws_d.cell(row, 7, c.get("dead") or 0)
        row += 1
    _autosize(ws_d)

    # ---------- Hoja 5: Medias + análisis ----------
    ws_m = wb.create_sheet("Análisis")
    _heading(ws_m, "A1", f"Análisis — {data['variable_label']}")
    _header_row(
        ws_m,
        3,
        ["Tratamiento", "Etiqueta", "n", "Media", "EE", "Letra", "Eficacia %"],
    )
    for i, m in enumerate(data["means"], start=4):
        ws_m.cell(i, 1, m["treatment"])
        ws_m.cell(i, 2, data["treatment_labels"].get(m["treatment"], ""))
        ws_m.cell(i, 3, m["n"])
        ws_m.cell(i, 4, round(m["mean"], 4))
        ws_m.cell(i, 5, round(m["se"], 4) if m["se"] is not None else None)
        ws_m.cell(i, 6, m["letter"])
        eff = m["efficacy_pct"]
        ws_m.cell(i, 7, round(eff, 2) if eff is not None else None)

    base = len(data["means"]) + 6
    ws_m.cell(base, 1, "ANOVA").font = Font(bold=True)
    _header_row(ws_m, base + 1, ["Fuente", "gl", "SC", "F", "p"])
    for i, a in enumerate(data["anova"], start=base + 2):
        ws_m.cell(i, 1, a["source"].replace("C(", "").replace(")", ""))
        ws_m.cell(i, 2, a["df"])
        ws_m.cell(i, 3, round(a["sum_sq"], 4) if a["sum_sq"] is not None else None)
        ws_m.cell(i, 4, round(a["F"], 3) if a["F"] is not None else None)
        ws_m.cell(i, 5, round(a["PR(>F)"], 5) if a["PR(>F)"] is not None else None)

    base2 = base + len(data["anova"]) + 4
    ws_m.cell(base2, 1, "Tukey HSD (α=0.05)").font = Font(bold=True)
    _header_row(ws_m, base2 + 1, ["A", "B", "Δ", "p ajustada", "Significativo"])
    for i, t in enumerate(data["tukey"], start=base2 + 2):
        ws_m.cell(i, 1, t["group1"])
        ws_m.cell(i, 2, t["group2"])
        ws_m.cell(i, 3, round(t["meandiff"], 4))
        ws_m.cell(i, 4, round(t["p_adj"], 5))
        ws_m.cell(i, 5, "sí" if t["reject"] else "")
    _autosize(ws_m)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =================== Helpers de formato ===================

def _heading(ws, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=14, color="2E7D32")


def _header_row(ws, row: int, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="F0F0F0")
    bold = Font(bold=True)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row, i, h)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left")


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)


# =================== Reporte consolidado de ensayo ===================


def _analyze_one_evaluation(
    evaluation: dict[str, Any],
    plots: list[dict[str, Any]],
    counts: list[dict[str, Any]],
    variable: str,
    control_key: str | None,
    pre_means: dict[str, float] | None,
) -> dict[str, Any] | None:
    """Corre ANOVA + Tukey + CLD + Abbott + (opcional) Henderson-Tilton para una evaluación.

    Devuelve None si no hay datos suficientes.
    """
    observations, _ = _build_dataset(plots, counts, variable)
    if len(observations) < 4:
        return None

    df = pd.DataFrame(observations)
    use_block = df["block"].notna().all() and df["block"].nunique() >= 2
    result = rcbd_anova_tukey(
        df,
        value_col="value",
        treatment_col="treatment",
        block_col="block" if use_block else None,
    )

    means_by_trt = {row["treatment"]: row["mean"] for row in result["means"]}
    differing = [(t["group1"], t["group2"]) for t in result["tukey"] if t["reject"]]
    letters = compact_letter_display(means_by_trt, differing)
    control_mean = means_by_trt.get(control_key) if control_key else None

    enriched: list[dict[str, Any]] = []
    for row in result["means"]:
        trt = row["treatment"]
        is_control = control_key is not None and trt == control_key

        abbott_eff = (
            0.0
            if is_control
            else (abbott(control_mean, row["mean"]) if control_mean is not None else None)
        )

        ht_eff: float | None = None
        if (
            pre_means is not None
            and control_key in pre_means
            and trt in pre_means
            and control_mean is not None
            and not is_control
        ):
            ht_eff = henderson_tilton(
                control_pre=pre_means[control_key],
                control_post=control_mean,
                treated_pre=pre_means[trt],
                treated_post=row["mean"],
            )

        enriched.append(
            {
                **row,
                "letter": letters.get(trt, ""),
                "abbott_pct": abbott_eff,
                "ht_pct": ht_eff,
            }
        )

    enriched.sort(key=lambda r: -r["mean"])

    return {
        "evaluation": evaluation,
        "means": enriched,
        "means_by_trt": means_by_trt,  # para próximas iteraciones HT
        "anova": result["anova"],
        "tukey": result["tukey"],
        "cv_pct": result["cv_pct"],
        "mse": result["mse"],
        "n_obs": result["n_obs"],
    }


def analyze_trial(trial_id: str, variable: str = TOTAL_VARIABLE) -> dict[str, Any]:
    """Análisis consolidado de un ensayo: ANOVA por evaluación, curva poblacional,
    Henderson-Tilton vs. el primer conteo (pre-aplicación)."""
    trial = db.get_trial(trial_id)
    if not trial:
        raise LookupError(f"Ensayo {trial_id} no encontrado")
    treatments = db.get_treatments(trial_id)
    plots = db.get_plots(trial_id)
    _hydrate_plots(plots, treatments)

    evaluations = db.list_evaluations_for_trial(trial_id)
    if not evaluations:
        raise ValueError("El ensayo no tiene evaluaciones cargadas.")

    control_key = None
    for t in treatments:
        if t["is_control"]:
            control_key = f"T{t['number']}"
            break

    per_eval: list[dict[str, Any]] = []
    pre_means: dict[str, float] | None = None
    for ev in evaluations:
        counts = db.get_pest_counts(ev["id"])
        analysis = _analyze_one_evaluation(
            ev, plots, counts, variable, control_key, pre_means
        )
        if analysis is None:
            continue
        per_eval.append(analysis)
        if pre_means is None:
            pre_means = analysis["means_by_trt"]

    if not per_eval:
        raise ValueError(
            "Ninguna evaluación tiene observaciones suficientes (≥2 tratamientos × ≥2 bloques)."
        )

    # Serie temporal por tratamiento para la curva poblacional
    time_series: dict[str, list[dict[str, Any]]] = {}
    for entry in per_eval:
        ev = entry["evaluation"]
        for m in entry["means"]:
            time_series.setdefault(m["treatment"], []).append(
                {
                    "evaluated_at": ev["evaluated_at"],
                    "dda": ev.get("days_after_application"),
                    "mean": m["mean"],
                    "se": m["se"],
                }
            )

    # Tabla resumen de eficacia: filas = tratamiento, columnas = DDA/fecha
    eff_matrix_rows: list[dict[str, Any]] = []
    for trt in sorted(time_series.keys()):
        row: dict[str, Any] = {"treatment": trt, "by_dda": []}
        for entry in per_eval:
            mean_row = next((m for m in entry["means"] if m["treatment"] == trt), None)
            if mean_row is None:
                row["by_dda"].append(None)
                continue
            row["by_dda"].append(
                {
                    "dda": entry["evaluation"].get("days_after_application"),
                    "evaluated_at": entry["evaluation"]["evaluated_at"],
                    "mean": mean_row["mean"],
                    "letter": mean_row["letter"],
                    "abbott_pct": mean_row["abbott_pct"],
                    "ht_pct": mean_row["ht_pct"],
                }
            )
        eff_matrix_rows.append(row)

    return {
        "trial": trial,
        "treatments": treatments,
        "plots": plots,
        "control_treatment": control_key,
        "variable": variable,
        "variable_label": _variable_label(variable),
        "per_evaluation": per_eval,
        "time_series": time_series,
        "efficacy_matrix": eff_matrix_rows,
    }


def population_curve_svg(
    time_series: dict[str, list[dict[str, Any]]],
    control_key: str | None,
    width: int = 520,
    height: int = 260,
) -> str:
    """SVG inline de la curva poblacional. Eje X = fecha, eje Y = media."""
    if not time_series:
        return ""

    # Recolectar todas las fechas únicas
    all_dates = sorted(
        {pt["evaluated_at"] for series in time_series.values() for pt in series}
    )
    all_values = [
        pt["mean"]
        for series in time_series.values()
        for pt in series
        if pt["mean"] is not None
    ]
    if not all_values:
        return ""

    max_v = max(all_values) * 1.15 if max(all_values) > 0 else 1.0

    pad_l, pad_r, pad_t, pad_b = 56, 16, 30, 56
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b

    def x_pos(date: str) -> float:
        idx = all_dates.index(date)
        if len(all_dates) == 1:
            return pad_l + plot_w / 2
        return pad_l + (idx / (len(all_dates) - 1)) * plot_w

    def y_pos(value: float) -> float:
        return pad_t + (1 - value / max_v) * plot_h

    treatments_sorted = sorted(
        time_series.keys(),
        key=lambda t: (0 if t == control_key else 1, int(t.lstrip("T") or "0")),
    )
    n = len(treatments_sorted)
    colors = [f"hsl({int(360 * i / max(n, 1))}, 60%, 45%)" for i in range(n)]

    parts: list[str] = [
        f'<svg viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg" '
        'style="font-family: -apple-system, Helvetica, sans-serif;">'
    ]

    # Ejes
    parts.append(
        f'<line x1="{pad_l}" y1="{pad_t}" x2="{pad_l}" y2="{pad_t + plot_h}" stroke="#888" stroke-width="1"/>'
    )
    parts.append(
        f'<line x1="{pad_l}" y1="{pad_t + plot_h}" x2="{pad_l + plot_w}" y2="{pad_t + plot_h}" stroke="#888" stroke-width="1"/>'
    )

    # Y ticks
    for tick in [0.0, max_v * 0.25, max_v * 0.5, max_v * 0.75, max_v]:
        y = y_pos(tick)
        parts.append(
            f'<line x1="{pad_l - 3}" y1="{y:.1f}" x2="{pad_l}" y2="{y:.1f}" stroke="#888"/>'
        )
        parts.append(
            f'<text x="{pad_l - 6}" y="{y + 3:.1f}" font-size="9" text-anchor="end" fill="#555">{tick:.1f}</text>'
        )
        parts.append(
            f'<line x1="{pad_l}" y1="{y:.1f}" x2="{pad_l + plot_w}" y2="{y:.1f}" stroke="#eee" stroke-width="1"/>'
        )

    # X labels (fechas) — rotadas
    for date in all_dates:
        x = x_pos(date)
        label = date[:10]
        parts.append(
            f'<text x="{x:.1f}" y="{pad_t + plot_h + 14:.1f}" font-size="9" text-anchor="middle" fill="#555">{label}</text>'
        )

    # Líneas + puntos por tratamiento
    for i, trt in enumerate(treatments_sorted):
        color = colors[i]
        series = time_series[trt]
        d_parts = []
        for j, pt in enumerate(series):
            cmd = "M" if j == 0 else "L"
            d_parts.append(f"{cmd}{x_pos(pt['evaluated_at']):.1f},{y_pos(pt['mean']):.1f}")
        parts.append(
            f'<path d="{" ".join(d_parts)}" stroke="{color}" stroke-width="2" fill="none"/>'
        )
        for pt in series:
            parts.append(
                f'<circle cx="{x_pos(pt["evaluated_at"]):.1f}" cy="{y_pos(pt["mean"]):.1f}" r="3" fill="{color}"/>'
            )

    # Leyenda en fila
    legend_y = height - 12
    legend_x = pad_l
    for i, trt in enumerate(treatments_sorted):
        color = colors[i]
        label = f"{trt}{' (testigo)' if trt == control_key else ''}"
        parts.append(
            f'<circle cx="{legend_x:.1f}" cy="{legend_y:.1f}" r="3" fill="{color}"/>'
        )
        parts.append(
            f'<text x="{legend_x + 7:.1f}" y="{legend_y + 3:.1f}" font-size="9" fill="#333">{label}</text>'
        )
        legend_x += 70 if trt != control_key else 90

    parts.append("</svg>")
    return "".join(parts)


def render_trial_pdf(trial_id: str, variable: str = TOTAL_VARIABLE) -> bytes:
    """Renderiza el PDF consolidado del ensayo."""
    from weasyprint import HTML  # type: ignore

    data = analyze_trial(trial_id, variable)
    trial = data["trial"]
    crop = db.get_crop(trial.get("crop_id"))
    client = db.get_client(trial.get("client_id"))
    pest = db.get_pest(trial.get("pest_id"))

    treatment_labels = {f"T{t['number']}": t.get("label", "") for t in data["treatments"]}
    curve_svg = population_curve_svg(data["time_series"], data["control_treatment"])

    html = _env.get_template("trial_report.html").render(
        trial=trial,
        treatments=data["treatments"],
        per_evaluation=data["per_evaluation"],
        efficacy_matrix=data["efficacy_matrix"],
        control=data["control_treatment"],
        treatment_labels=treatment_labels,
        variable_label=data["variable_label"],
        curve_svg=curve_svg,
        crop=crop,
        client=client,
        pest=pest,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return HTML(string=html).write_pdf()


def render_trial_xlsx(trial_id: str, variable: str = TOTAL_VARIABLE) -> bytes:
    """Excel consolidado del ensayo (Resumen, Tratamientos, Parcelas, Conteos
    crudos de todas las evaluaciones, Eficacia consolidada y una hoja por evaluación)."""
    data = analyze_trial(trial_id, variable)
    trial = data["trial"]
    treatments = data["treatments"]
    plots = data["plots"]

    wb = Workbook()

    # ---------- Hoja 1: Resumen ----------
    ws = wb.active
    ws.title = "Resumen"
    _heading(ws, "A1", f"{trial['code']} — {trial['name']}")
    rows = [
        ("Código", trial["code"]),
        ("Nombre", trial["name"]),
        ("Tipo", trial.get("trial_type")),
        ("Diseño", trial.get("design")),
        ("# Tratamientos", trial.get("n_treatments")),
        ("# Repeticiones", trial.get("n_replicates")),
        ("Ubicación", trial.get("location") or ""),
        ("Variable analizada", data["variable_label"]),
        ("Evaluaciones incluidas", len(data["per_evaluation"])),
        ("Testigo", data["control_treatment"] or "—"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    _autosize(ws)

    # ---------- Hoja 2: Tratamientos ----------
    ws_t = wb.create_sheet("Tratamientos")
    _header_row(ws_t, 1, ["#", "Etiqueta", "Testigo"])
    for i, t in enumerate(treatments, start=2):
        ws_t.cell(i, 1, f"T{t['number']}")
        ws_t.cell(i, 2, t.get("label"))
        ws_t.cell(i, 3, "sí" if t.get("is_control") else "")
    _autosize(ws_t)

    # ---------- Hoja 3: Parcelas ----------
    ws_p = wb.create_sheet("Parcelas")
    _header_row(ws_p, 1, ["Parcela ID", "Bloque", "Col", "Tratamiento", "Etiqueta"])
    for i, p in enumerate(plots, start=2):
        ws_p.cell(i, 1, p["id"])
        ws_p.cell(i, 2, p["block"])
        ws_p.cell(i, 3, p.get("position_col"))
        ws_p.cell(i, 4, f"T{p['_treatment_number']}")
        ws_p.cell(i, 5, p.get("_treatment_label"))
    _autosize(ws_p)

    # ---------- Hoja 4: Conteos crudos (todas las evaluaciones) ----------
    ws_c = wb.create_sheet("Conteos")
    _header_row(
        ws_c,
        1,
        [
            "Evaluación",
            "Fecha",
            "DDA",
            "Bloque",
            "Col",
            "Tratamiento",
            "Muestra",
            "Estadio",
            "Vivos",
            "Muertos",
        ],
    )
    plot_by_id = {p["id"]: p for p in plots}
    row_idx = 2
    for i_ev, entry in enumerate(data["per_evaluation"], start=1):
        ev = entry["evaluation"]
        counts = db.get_pest_counts(ev["id"])
        for c in counts:
            plot = plot_by_id.get(c["plot_id"])
            if not plot:
                continue
            ws_c.cell(row_idx, 1, f"E{i_ev}")
            ws_c.cell(row_idx, 2, ev["evaluated_at"])
            ws_c.cell(row_idx, 3, ev.get("days_after_application"))
            ws_c.cell(row_idx, 4, plot["block"])
            ws_c.cell(row_idx, 5, plot.get("position_col"))
            ws_c.cell(row_idx, 6, f"T{plot['_treatment_number']}")
            ws_c.cell(row_idx, 7, c.get("sample_index"))
            ws_c.cell(
                row_idx, 8, STAGE_LABELS.get(c.get("life_stage") or "", c.get("life_stage"))
            )
            ws_c.cell(row_idx, 9, c.get("alive") or 0)
            ws_c.cell(row_idx, 10, c.get("dead") or 0)
            row_idx += 1
    _autosize(ws_c)

    # ---------- Hoja 5: Eficacia consolidada ----------
    ws_e = wb.create_sheet("Eficacia")
    header = ["Tratamiento"]
    for entry in data["per_evaluation"]:
        ev = entry["evaluation"]
        dda = ev.get("days_after_application")
        suffix = f"DDA {dda}" if dda is not None else ev["evaluated_at"][:10]
        header.append(f"{suffix} HT%")
        header.append(f"{suffix} Abbott%")
    _header_row(ws_e, 1, header)
    for i, row in enumerate(data["efficacy_matrix"], start=2):
        ws_e.cell(i, 1, row["treatment"])
        col = 2
        for cell in row["by_dda"]:
            if cell is None:
                ws_e.cell(i, col, None)
                ws_e.cell(i, col + 1, None)
            else:
                ws_e.cell(
                    i,
                    col,
                    round(cell["ht_pct"], 2) if cell.get("ht_pct") is not None else None,
                )
                ws_e.cell(
                    i,
                    col + 1,
                    round(cell["abbott_pct"], 2)
                    if cell.get("abbott_pct") is not None
                    else None,
                )
            col += 2
    _autosize(ws_e)

    # ---------- Una hoja por evaluación con medias + ANOVA ----------
    for i_ev, entry in enumerate(data["per_evaluation"], start=1):
        ev = entry["evaluation"]
        sheet_name = f"E{i_ev}"
        ws_ev = wb.create_sheet(sheet_name[:31])
        _heading(
            ws_ev,
            "A1",
            f"Evaluación {i_ev} — {ev['evaluated_at'][:16]}"
            + (f" · DDA {ev['days_after_application']}" if ev.get("days_after_application") is not None else ""),
        )
        _header_row(
            ws_ev,
            3,
            ["Trat", "n", "Media", "EE", "Letra", "Abbott%", "HT%"],
        )
        for j, m in enumerate(entry["means"], start=4):
            ws_ev.cell(j, 1, m["treatment"])
            ws_ev.cell(j, 2, m["n"])
            ws_ev.cell(j, 3, round(m["mean"], 4))
            ws_ev.cell(j, 4, round(m["se"], 4) if m["se"] is not None else None)
            ws_ev.cell(j, 5, m["letter"])
            ws_ev.cell(
                j, 6, round(m["abbott_pct"], 2) if m["abbott_pct"] is not None else None
            )
            ws_ev.cell(
                j, 7, round(m["ht_pct"], 2) if m["ht_pct"] is not None else None
            )

        base = len(entry["means"]) + 6
        ws_ev.cell(base, 1, "ANOVA").font = Font(bold=True)
        _header_row(ws_ev, base + 1, ["Fuente", "gl", "SC", "F", "p"])
        for k, a in enumerate(entry["anova"], start=base + 2):
            ws_ev.cell(k, 1, a["source"].replace("C(", "").replace(")", ""))
            ws_ev.cell(k, 2, a["df"])
            ws_ev.cell(k, 3, round(a["sum_sq"], 4) if a["sum_sq"] is not None else None)
            ws_ev.cell(k, 4, round(a["F"], 3) if a["F"] is not None else None)
            ws_ev.cell(
                k, 5, round(a["PR(>F)"], 5) if a["PR(>F)"] is not None else None
            )
        _autosize(ws_ev)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
